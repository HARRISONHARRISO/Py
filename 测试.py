# -*- coding: utf-8 -*-
"""
Created on Tue Dec 28 09:48:27 2021

@author: zzy
"""
from matplotlib.backends.backend_qt5 import NavigationToolbar2QT
from matplotlib.figure import Figure
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from PyQt5.QtWidgets import QMainWindow, QApplication, QWidget, QGridLayout
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QApplication, QMainWindow,QMessageBox,QTableWidgetItem,QFileDialog
import os
import pandas as pd 
import numpy as np
import math
import openpyxl 
import sys

import random 
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtCore import *

#from upipe import Ui_Form2
from qcl import Ui_Form2
#from untitled import Ui_Form
from untitled2 import Ui_Form
#from postprocessing_app import Ui_Form3
from post import Ui_Form3
class MyMainForm(QMainWindow, Ui_Form):
    def __init__(self, parent=None):
        super(MyMainForm, self).__init__(parent)
        self.setupUi(self)
    def chushi(self):
        self.show()
    def open_qianchuli(self):
        main.pushButton.clicked.disconnect()
        if main.radioButton.isChecked()==True :
            main.pushButton.clicked.connect(child.open_q)
            
    def open_houchuli(self):
        main.pushButton.clicked.disconnect() 
        if  main.radioButton_2.isChecked()==True:
            main.pushButton.clicked.connect(child2.open_h)
           
class child(QMainWindow, Ui_Form2):
    def __init__(self, parent=None):
        super(child, self).__init__(parent)
        self.setupUi(self)
        self.num=0
        self.pushButton.clicked.connect(self.openfile_1)
        self.pushButton_2.clicked.connect(self.openfile_2)
        ####背景图设置
        self.setWindowTitle('OLGA前处理工具')
#        palette = QPalette()
#        palette.setBrush(QPalette.Background,QBrush(QPixmap('2.jpg')))
#        self.setPalette(palette)
######
       
    def open_q(self):
        self.show()    
    def openfile_1(self):
        self.textBrowser.clear()
        fileName,fileType=QtWidgets.QFileDialog.getOpenFileName(self,"选取文件",os.getcwd(),
                                                                "All Files(*);;Text Files(*.txt)")
        self.textBrowser.insertPlainText(fileName)
    def openfile_2(self):
        self.textBrowser_2.clear()
        fileName,fileType=QtWidgets.QFileDialog.getOpenFileName(self,"选取文件",os.getcwd(),
                                                                "All Files(*);;Text Files(*.txt)")
        self.textBrowser_2.insertPlainText(fileName) 
    #管道划分函数    
    def huafen(self):
        
        ts1=self.lineEdit.text()
        if len(ts1)==0:
           QMessageBox.about(self,"提示","请输入最大管段长度")
        ts2=self.textBrowser.toPlainText()
        if len(ts2)==0:
           QMessageBox.about(self,"提示","请选择tab文件") 
        ts3=self.textBrowser_2.toPlainText()
        if len(ts3)==0:
           QMessageBox.about(self,"提示","请选择excel文件")
        
        
        fileName_excel=self.textBrowser_2.toPlainText()
#        f= "数据录入.xlsx"
        dataset=pd.read_excel(fileName_excel)
        wb=openpyxl.Workbook()
        ws=wb.active
        # dataset=pd.read_excel(f,sheet_name='Sheet1')

        #        fileName2=self.textBrowser_2.toPlainText()

        def AA():
            
            jihe1=[]
            jihe2=[]            
            num=1
            df_excel=pd.read_excel(fileName_excel)
            # df_excel=pd.read_excel(f)
            for m in range(0,df_excel.shape[0]):
                
                
                for j in range(0,5):
                    jihe1.append(dataset.iloc[m,j+6])
                    jihe2.append(float(self.lineEdit.text()))
            
            
             
                for k in range(0,len(jihe1)):
                    duanshu=jihe1[k]/jihe2[k]
                    duanshu2=math.ceil(duanshu)
                    if duanshu2<2:
                        duanshu2=2
                    junfen=jihe1[k]/duanshu2
                    a=[]
                    g=[]
                    m=0
                    for i in range(0,duanshu2):   
                        m=m+junfen
                        b=m/jihe1[k]
                        a.append(m)
                        g.append(b)
                    c=[]
                    d=[]
                    for i in range (0,len(g)-1):#原始为len(g),然后把测试内容删除
                        # jisuanzhi=-3.53*pow(g[i],5)+8.386*pow(g[i],4)-7.055*pow(g[i],3)+2.484*pow(g[i],2)+0.7183*(g[i])-0.003213
                        jisuanzhi=1.08199627*pow(g[i],6)-6.43186814*pow(g[i],5)+12.1201047*pow(g[i],4)-10.2030968*pow(g[i],3)+4.12417864*pow(g[i],2)+0.309604037*(g[i])+0.000582141097
                        c.append(jisuanzhi)
                        fanguiyi=jisuanzhi*jihe1[k]
                        d.append(fanguiyi)
                        #####测试
                       
                    fanguiyi_=jihe1[k]
                    d.append(fanguiyi_)
                        
                        
                        
                        
                        
                        
                        
                        
                        #####
                    e=[]
                    for i in range (0,len(g)-1):
                        shiji=d[i+1]-d[i]
                        e.append(shiji)
                    e.insert(0,d[0]) 
                    for i in range(0, len(e)):
                        ws.cell(row=i+1,column=num).value=e[i]
                    num=num+1
                jihe1=[]
                jihe2=[]            
                        
                       
            # wb.save('划分结果_实验对比.xlsx') 
            wb.save('划分结果_实验.xlsx')  
            
            
            
            df_normal=pd.read_excel('划分结果_实验.xlsx',header=None)
            df2=pd.read_excel('划分结果_实验.xlsx',header=None)
            ####第一列
            for i in range(0,df_excel.shape[0]*5,5):
                shanchu=df2.dropna(axis=0, how='any' ,inplace=False,subset=[i])
                shanchu_h=df2.dropna(axis=0, how='any' ,inplace=False,subset=[i+1])
                lieshu_q=shanchu.shape[0]
                lieshu_h=shanchu_h.shape[0]
                qian_2=df2.iloc[lieshu_q-1,i]
                hou_2=df2.iloc[0,i+1]
                if qian_2>2*hou_2:
                   # df2.iloc[0,i]=2*df2.iloc[lieshu_q-1,i-1]
            
                   df2.iloc[lieshu_q-1,i]=random.uniform(df2.iloc[0,i+1],2*df2.iloc[0,i+1])
                   # print(df2.iloc[0,i])
                num_kongzhi=1
                ngx=[]
                gx=[]
                for j in range(lieshu_q-1,-1,-1):
                    if 2*df2.iloc[j,i]<df2.iloc[j-1,i]:
                       ngx.append(df2.iloc[j-1,i])
                       # df2.iloc[j,i]=2*df2.iloc[j-1,i]
            
                       df2.iloc[j-1,i]=random.uniform(df2.iloc[j,i],2*df2.iloc[j,i])
                       num_kongzhi=num_kongzhi+1
                       gx.append(df2.iloc[j-1,i])
                    else:
                        break
                # print(num_kongzhi)
                ngx_sum=sum(ngx)
                gx_sum=sum(gx)
                chazhi=(ngx_sum+df_normal.iloc[lieshu_q-1,i])-(gx_sum+df2.iloc[lieshu_q-1,i])
                if num_kongzhi-1==0:
                    df2.iloc[0,i]=df2.iloc[0,i]+chazhi
                else:
                    for k in range(lieshu_q-num_kongzhi-1,-1,-1):
                        df2.iloc[k,i]=df2.iloc[k,i]+chazhi/(lieshu_q-num_kongzhi)        
             
            ######中间列
            for w in range(1,4):
                for i in range(w,df_excel.shape[0]*5+w,5):
                    shanchu_i=df2.dropna(axis=0, how='any' ,inplace=False,subset=[i-1])
                    shanchu=df2.dropna(axis=0, how='any' ,inplace=False,subset=[i])
                    shanchu_h=df2.dropna(axis=0, how='any' ,inplace=False,subset=[i+1])
                    lieshu_i=shanchu_i.shape[0]
                    lieshu_q=shanchu.shape[0]
                    lieshu_h=shanchu_h.shape[0]
                    qian_i=df2.iloc[lieshu_i-1,i-1]
                    hou_i=df2.iloc[0,i]
                    qian_2=df2.iloc[lieshu_q-1,i]
                    hou_2=df2.iloc[0,i+1]
                    
                    
                    
                    ########
                    if qian_2>2*hou_2 and 2*qian_i<hou_i:
                        df2.iloc[lieshu_q-1,i]=random.uniform(df2.iloc[0,i+1],2*df2.iloc[0,i+1])
                        df2.iloc[0,i]=random.uniform(df2.iloc[lieshu_i-1,i-1],2*df2.iloc[lieshu_i-1,i-1])
                        num_kongzhi_1=1
                        num_kongzhi_2=0
                        ngx_1=[]
                        ngx_2=[]
                        gx_1=[]
                        gx_2=[]
                        
                        for j in range(lieshu_q-1,-1,-1):
                            if 2*df2.iloc[j,i]<df2.iloc[j-1,i]:
                              ngx_1.append(df2.iloc[j-1,i])
                              # df2.iloc[j,i]=2*df2.iloc[j-1,i]
            
                              df2.iloc[j-1,i]=random.uniform(df2.iloc[j,i],2*df2.iloc[j,i])
                              num_kongzhi_1=num_kongzhi_1+1
                              gx_1.append(df2.iloc[j-1,i])
                            else:
                                break
                        for j in range(1,lieshu_q):
                            if df2.iloc[j,i]>2*df2.iloc[j-1,i]:
                              ngx_2.append(df2.iloc[j,i])
                              # df2.iloc[j,i]=2*df2.iloc[j-1,i]
            
                              df2.iloc[j,i]=random.uniform(df2.iloc[j-1,i],2*df2.iloc[j-1,i])
                              num_kongzhi_2=j
                              gx_2.append(df2.iloc[j,i])
                            else:
                                break
                        ngx_sum_1=sum(ngx_1)
                        ngx_sum_2=sum(ngx_2)
                        ngx_sum=ngx_sum_1+ngx_sum_2
                        gx_sum_1=sum(gx_1)
                        gx_sum_2=sum(gx_2)
                        gx_sum=gx_sum_1+gx_sum_2
                        kongzhi=num_kongzhi_1+num_kongzhi_2
                        # print(df_normal.iloc[lieshu_q-1,i])
                        chazhi=(ngx_sum+df_normal.iloc[lieshu_q-1,i]+df_normal.iloc[0,i])-(gx_sum+df2.iloc[lieshu_q-1,i]+df2.iloc[0,i])
                        # print(gx_sum_1,gx_sum_2,chazhi)
                        fsdaf=[]
                        for k in range(num_kongzhi_2+1,lieshu_q-num_kongzhi_1):
                            # fsdaf.append( df2.iloc[k,i])
                            # print(kongzhi)
                            df2.iloc[k,i]=df2.iloc[k,i]+chazhi/(lieshu_q-kongzhi-1)
                            
                       
                       
                       
                       
                       
                    #####
                    # elif qian_2>2*hou_2 and 0.5*qian_i<=hou_i<=2*qian_i:
                    elif qian_2>2*hou_2 and hou_i<=2*qian_i:    
                        # df2.iloc[0,i]=2*df2.iloc[lieshu_q-1,i-1]
            
                        df2.iloc[lieshu_q-1,i]=random.uniform(df2.iloc[0,i+1],2*df2.iloc[0,i+1])
                        # print(df2.iloc[0,i])
                        num_kongzhi=1
                        ngx=[]
                        gx=[]
                        for j in range(lieshu_q-1,-1,-1):
                            if 2*df2.iloc[j,i]<df2.iloc[j-1,i]:
                                ngx.append(df2.iloc[j-1,i])
                                # df2.iloc[j,i]=2*df2.iloc[j-1,i]
                
                                df2.iloc[j-1,i]=random.uniform(df2.iloc[j,i],2*df2.iloc[j,i])
                                num_kongzhi=num_kongzhi+1
                                gx.append(df2.iloc[j-1,i])
                            else:
                                break
                        # print(num_kongzhi)
                        ngx_sum=sum(ngx)
                        gx_sum=sum(gx)
                        chazhi=(ngx_sum+df_normal.iloc[lieshu_q-1,i])-(gx_sum+df2.iloc[lieshu_q-1,i])
                        if num_kongzhi-1==0:
                            df2.iloc[0,i]=df2.iloc[0,i]+chazhi
                        else:
                            for k in range(lieshu_q-num_kongzhi-1,-1,-1):
                                df2.iloc[k,i]=df2.iloc[k,i]+chazhi/(lieshu_q-num_kongzhi)
                
                #####
                    # elif 2*qian_i<hou_i and 0.5*hou_2<=qian_2<=2*hou_2:
                    elif 2*qian_i<hou_i and qian_2<=hou_2:   
                        # df2.iloc[0,i]=2*df2.iloc[lieshu_q-1,i-1]
            
                        df2.iloc[0,i]=random.uniform(df2.iloc[lieshu_i-1,i-1],2*df2.iloc[lieshu_i-1,i-1])
                        # print(df2.iloc[0,i])
                        num_kongzhi=0
                        ngx=[]
                        gx=[]
                        for j in range(1,lieshu_q):
                            if df2.iloc[j,i]>2*df2.iloc[j-1,i]:
                                ngx.append(df2.iloc[j,i])
                                # df2.iloc[j,i]=2*df2.iloc[j-1,i]
                
                                df2.iloc[j,i]=random.uniform(df2.iloc[j-1,i],2*df2.iloc[j-1,i])
                                num_kongzhi=j
                                gx.append(df2.iloc[j,i])
                            else:
                                break
                        ngx_sum=sum(ngx)
                        gx_sum=sum(gx)
                        chazhi=(ngx_sum+df_normal.iloc[0,i])-(gx_sum+df2.iloc[0,i])
                        if num_kongzhi+1==lieshu_q:
                            df2.iloc[lieshu_q,i]=df2.iloc[lieshu_q,i]+chazhi
                        else:
                            for k in range(num_kongzhi+1,lieshu_q):
                                df2.iloc[k,i]=df2.iloc[k,i]+chazhi/(lieshu_q-num_kongzhi-1)
            #   ######## 更新1
            for k in range(0,df_excel.shape[0]):
                  
                  for i in range(2+k*5,3+k*5):
                      
                      shanchu_hh=df2.dropna(axis=0, how='any' ,inplace=False,subset=[i])
                      lieshu_hh=shanchu_hh.shape[0]
                      sum_chazhi=0
                      for s in range(lieshu_hh-1,0,-1):
                          if  2*df2.iloc[s,i]<df2.iloc[s-1,i]:
                              sd=random.uniform(df2.iloc[s,i],2*df2.iloc[s,i])
                              sum_chazhi=df2.iloc[s-1,i]-sd+sum_chazhi
                              df2.iloc[s-1,i]=sd   
                          else:
                              break
                      pingjun=sum_chazhi/(lieshu_hh-2)
                      for e in range(1,lieshu_hh-1):
                          df2.iloc[e,i]=df2.iloc[e,i]+pingjun
                           






               
            ######## 更新2   
            for k in range(0,df_excel.shape[0]):
                
                for i in range(1+k*5,2+k*5):
                    jgx=[]
                    shanchu_jgx=df2.dropna(axis=0, how='any' ,inplace=False,subset=[i])
                    shanchu_jgx_q=df2.dropna(axis=0, how='any' ,inplace=False,subset=[i-1])
                    shanchu_jgx_h=df2.dropna(axis=0, how='any' ,inplace=False,subset=[i+1])
                    
                    lieshu_jgx=shanchu_jgx.shape[0]
                    lieshu_jgx_q=shanchu_jgx_q.shape[0]
                    lieshu_jgx_h=shanchu_jgx_h.shape[0]
                    
                    for j in range(0,lieshu_jgx):
                        jgx.append(df2.iloc[j,i])
                        sum_jgx=sum(jgx)
                    if sum_jgx<dataset.iloc[k,7]:
                        kt=random.uniform(df2.iloc[lieshu_jgx_q-1,i-1],1.5*df2.iloc[lieshu_jgx_q-1,i-1])
                        mw=random.uniform(0.5*df2.iloc[0,i+1], 0.6*df2.iloc[0,i+1])
                        gx_ds=dataset.iloc[k,7]//(kt+mw)
                        gx_ds=int(gx_ds)
                        yushu=dataset.iloc[k,7]%(kt+mw)
                        if yushu>2*kt:
                            yushu1=random.uniform(kt, 2*kt)
                            yushu2=yushu-yushu1
                            for d in range(0,gx_ds):
                                df2.iloc[d,i]=kt
                            df2.iloc[gx_ds,i]=yushu1
                            df2.iloc[gx_ds+1,i]=yushu2
                            for h in range(0,gx_ds):
                                df2.iloc[gx_ds+h+2,i]=mw
                        elif yushu>2*mw:
                              yushu2=random.uniform(mw, 2*mw)
                              yushu1=yushu-yushu2
                              for d in range(0,gx_ds):
                                  df2.iloc[d,i]=kt
                              df2.iloc[gx_ds,i]=yushu1
                              df2.iloc[gx_ds+1,i]=yushu2
                              for h in range(0,gx_ds):
                                  df2.iloc[gx_ds+h+2,i]=mw
                        else:
                            for d in range(0,gx_ds):
                                df2.iloc[d,i]=kt
                            df2.iloc[gx_ds,i]=yushu
                            for h in range(0,gx_ds):
                                df2.iloc[gx_ds+h+1,i]=mw
                            
            
            
            
                for i in range(3+k*5,4+k*5):
                    jgx=[]
                    shanchu_jgx=df2.dropna(axis=0, how='any' ,inplace=False,subset=[i])
                    shanchu_jgx_q=df2.dropna(axis=0, how='any' ,inplace=False,subset=[i-1])
                    shanchu_jgx_h=df2.dropna(axis=0, how='any' ,inplace=False,subset=[i+1])
                    
                    lieshu_jgx=shanchu_jgx.shape[0]
                    lieshu_jgx_q=shanchu_jgx_q.shape[0]
                    lieshu_jgx_h=shanchu_jgx_h.shape[0]
                    
                    for j in range(0,lieshu_jgx):
                        jgx.append(df2.iloc[j,i])
                        sum_jgx=sum(jgx)
                    if sum_jgx<dataset.iloc[k,9]:
                        kt=random.uniform(0.5*df2.iloc[lieshu_jgx_q-1,i-1],0.6*df2.iloc[lieshu_jgx_q-1,i-1])
                        mw=random.uniform(df2.iloc[0,i+1], 1.5*df2.iloc[0,i+1])
                        gx_ds=dataset.iloc[k,9]//(kt+mw)
                        gx_ds=int(gx_ds)
                        yushu=dataset.iloc[k,9]%(kt+mw)
                        if yushu>2*kt:
                            yushu1=random.uniform(kt, 2*kt)
                            yushu2=yushu-yushu1
                            for d in range(0,gx_ds):
                                df2.iloc[d,i]=kt
                            df2.iloc[gx_ds,i]=yushu1
                            df2.iloc[gx_ds+1,i]=yushu2
                            for h in range(0,gx_ds):
                                df2.iloc[gx_ds+h+2,i]=mw
                        elif yushu>2*mw:
                              yushu2=random.uniform(mw, 2*mw)
                              yushu1=yushu-yushu2
                              for d in range(0,gx_ds):
                                  df2.iloc[d,i]=kt
                              df2.iloc[gx_ds,i]=yushu1
                              df2.iloc[gx_ds+1,i]=yushu2
                              for h in range(0,gx_ds):
                                  df2.iloc[gx_ds+h+2,i]=mw
                        else:
                            for d in range(0,gx_ds):
                                df2.iloc[d,i]=kt
                            df2.iloc[gx_ds,i]=yushu
                            for h in range(0,gx_ds):
                                df2.iloc[gx_ds+h+1,i]=mw
            
            
            
            
            
            
            
            
            
            
            
            
            
            
            
            
                 
            ####### 
            ####末列
            for i in range(4,df_excel.shape[0]*5+4,5):
                shanchu=df2.dropna(axis=0, how='any' ,inplace=False,subset=[i-1])
                shanchu_h=df2.dropna(axis=0, how='any' ,inplace=False,subset=[i])
                lieshu_q=shanchu.shape[0]
                lieshu_h=shanchu_h.shape[0]
                qian_2=df2.iloc[lieshu_q-1,i-1]
                hou_2=df2.iloc[0,i]
                if 2*qian_2<hou_2:
                    # df2.iloc[0,i]=2*df2.iloc[lieshu_q-1,i-1]
            
                    df2.iloc[0,i]=random.uniform(df2.iloc[lieshu_q-1,i-1],2*df2.iloc[lieshu_q-1,i-1])
                    # print(df2.iloc[0,i])
                num_kongzhi=0
                ngx=[]
                gx=[]
                for j in range(1,lieshu_h):
                    if df2.iloc[j,i]>2*df2.iloc[j-1,i]:
                        ngx.append(df2.iloc[j,i])
                        # df2.iloc[j,i]=2*df2.iloc[j-1,i]
            
                        df2.iloc[j,i]=random.uniform(df2.iloc[j-1,i],2*df2.iloc[j-1,i])
                        num_kongzhi=j
                        gx.append(df2.iloc[j,i])
                    else:
                        break
                ngx_sum=sum(ngx)
                gx_sum=sum(gx)
                chazhi=(ngx_sum+df_normal.iloc[0,i])-(gx_sum+df2.iloc[0,i])
                if num_kongzhi+1==lieshu_h:
                    df2.iloc[lieshu_h,i]=df2.iloc[lieshu_h,i]+chazhi
                else:
                    for k in range(num_kongzhi+1,lieshu_h):
                        df2.iloc[k,i]=df2.iloc[k,i]+chazhi/(lieshu_h-num_kongzhi-1)    
            return df2


        # a=AA()
        # b=AA()
        #####
        dd=[]
        df_excel=pd.read_excel(fileName_excel)
    
        a=AA()
        b=a
        # for m in range(0,df_excel.shape[0]):
        for m in range(0,1):   
           

            for n in range(0,1000):
                c=AA()
           
            # for m in range(0,1):   
                for i in range(0+5*m,5+5*m):
                    shanchu=c.dropna(axis=0, how='any' ,inplace=False,subset=[i])
                    for j in range(0,shanchu.shape[0]):
                        
                        # if a.iloc[j,i]!=np.nan:
                        dd.append(shanchu.iloc[j,i])
                num=0
                for k in range(len(dd)-1):
                    if  dd[k+1]/dd[k]>2 or dd[k+1]/dd[k]<0.5:
                        dd=[]
                        
                        break
                    else: 
                        num=num+1
                if num==len(dd)-1:
                    for i in range(0+5*m,5+5*m):    
                        for j in range(0,a.shape[0]):
                            b.iloc[j,i]=c.iloc[j,i]
                    dd=[]        
                    break                    
            # print(n)   
        for i in range(1,df_excel.shape[0]):
            for j in range(0,b.shape[0]):       
                b.iloc[j,i*5]=b.iloc[j,0]
                b.iloc[j,1+i*5]=b.iloc[j,1]
                b.iloc[j,2+i*5]=b.iloc[j,2]  
                b.iloc[j,3+i*5]=b.iloc[j,3]
                b.iloc[j,4+i*5]=b.iloc[j,4]
        b.to_excel('输出.xlsx',index=False,header=None)     
        
        
        
        
        
        
        QMessageBox.about(self,"提示","划分完毕!") 
      
        
        
        
        
    def rewrite(self):
         fileName_tab=self.textBrowser.toPlainText()
         fileName_excel=self.textBrowser_2.toPlainText() 
         
         # df_hf=pd.read_excel('输出.xlsx',index=False,header=None)
         df=pd.read_excel(fileName_excel)
         for i in range(0,df.shape[0]):
             f=open('example.opi','r',encoding='utf-8')
             newname=df.iloc[i,0]
             lines=f.readlines()
             alldata=[]
             for lines in lines:
                 alldata.append(lines)
                 
             fileName_split=fileName_tab.split('/')
             fileName_split_last=fileName_split[-1][-4:]#提取.tab关键词
             fileName_split_last=df.iloc[i,15]+fileName_split_last
             fileName_split.pop(-1)
             fileName_split.append(fileName_split_last)
             fileName_split = '/'.join(fileName_split)
             df_hf=pd.read_excel('输出.xlsx',header=None)   
             f=open('example.opi','r',encoding='utf-8')   
             data=[] 
             lines=f.readlines() 
             for lines in lines:
                 data.append(lines)
                 if '<Key Name="PVTFILE">' in lines:
                     break   

                #基础部分写入到新的opi中
             with open('%s.opi'%newname,'w')as f:
                 for u in range(len(data)):
                     f.writelines(data[u])
             with open('%s.opi'%newname,'a')as f:
                  f.writelines('        <Values>\n\
          <Value>%s</Value>\n\
        </Values>\n\
        <Unit />\n\
        <DefaultUnit>NoUnit</DefaultUnit>\n\
      </Key>\n\
    </KeyCollection>\n\
  </Keyword>\n\
  <Keyword>\n'%(fileName_split))
  #####
                  # path.split('/')[-1][:-4]
  
  ####
                  f.writelines(alldata[3002:3140])
                  f.writelines('          <Key Name="TEMPERATURE">\n\
            <Values>\n\
              <Value>%.2f</Value>\n\
            </Values>\n\
            <Unit>C</Unit>\n\
            <DefaultUnit>C</DefaultUnit>\n\
          </Key>\n\
          <Key Name="PRESSURE">\n\
            <Values>\n\
              <Value>%.2f</Value>\n\
            </Values>\n\
            <Unit>bara</Unit>\n\
            <DefaultUnit>Pa</DefaultUnit>\n\
          </Key>\n\
          <Key Name="FLUID">\n\
           <Values>\n\
            <Value>%s</Value>\n'%(df.iloc[i,3],df.iloc[i,4]/100,df.iloc[i,14]))#出口参数设置
                  f.writelines(alldata[3157:3186])
                  f.writelines('          <Key Name="FLUID">\n\
            <Values>\n\
              <Value>%s</Value>\n'%(df.iloc[i,14]))
                  f.writelines(alldata[3189:3209]) 
                  
                  
                  #第一根管线参数
                  b_1=0

                  for frist_line_h in range(0,df_hf.shape[0]):
                          if np.isnan(df_hf.iloc[frist_line_h,5*i])==False:
                              b_1=b_1+1 
                          else:
                              b_1=b_1
                  f.writelines('        <Tag>FLOWPATH_1.PIPE_4</Tag>\n\
        <Type>PIPE</Type>\n\
        <KeyCollection>\n\
          <Key Name="ROUGHNESS">\n\
            <Values>\n\
              <Value>%.2f</Value>\n\
            </Values>\n\
            <Unit>mm</Unit>\n\
            <DefaultUnit>m</DefaultUnit>\n\
          </Key>\n\
          <Key Name="LABEL">\n\
            <Values>\n\
              <Value>PIPE-1</Value>\n\
            </Values>\n\
            <Unit />\n\
            <DefaultUnit>NoUnit</DefaultUnit>\n\
          </Key>\n\
          <Key Name="NSEGMENT">\n\
            <Values>\n\
              <Value>%d</Value>\n\
            </Values>\n\
            <Unit />\n\
            <DefaultUnit>NoUnit</DefaultUnit>\n\
          </Key>\n\
          <Key Name="LSEGMENT">\n\
            <Values>\n'%(df.iloc[i,5],b_1))
                  for m in range(0,b_1):
                      f.writelines('              <Value>%.4f</Value>\n'%df_hf.iloc[m,5*i])
                  f.writelines('            </Values>\n\
            <Unit>m</Unit>\n\
            <DefaultUnit>m</DefaultUnit>\n\
          </Key>\n\
          <Key Name="LENGTH">\n\
            <Values>\n\
              <Value>%.2f</Value>\n\
            </Values>\n\
            <Unit>m</Unit>\n\
            <DefaultUnit>m</DefaultUnit>\n\
          </Key>\n\
          <Key Name="ELEVATION">\n\
            <Values>\n\
              <Value>0</Value>\n\
            </Values>\n\
            <Unit>m</Unit>\n\
            <DefaultUnit>m</DefaultUnit>\n\
          </Key>\n\
          <Key Name="DIAMETER">\n\
            <Values>\n\
              <Value>%.2f</Value>\n\
            </Values>\n\
            <Unit>mm</Unit>\n\
            <DefaultUnit>m</DefaultUnit>\n\
          </Key>\n\
        </KeyCollection>\n\
      </Keyword>\n\
      <Keyword>\n'%(df.iloc[i,6],df.iloc[i,11]))
                  f.writelines(alldata[3267:3377]) 
                  f.writelines('          <Key Name="TEMPERATURE">\n\
            <Values>\n\
              <Value>%.2f</Value>\n\
            </Values>\n\
            <Unit>C</Unit>\n\
            <DefaultUnit>C</DefaultUnit>\n\
          </Key>\n\
          <Key Name="STDFLOWRATE">\n\
            <Values>\n\
              <Value>%.2f</Value>\n\
            </Values>\n'%(df.iloc[i,1],df.iloc[i,2]*10000))
                  f.writelines(alldata[3388:3404])
                  b_2=0
                  for sec_line_h in range(0,df_hf.shape[0]):
                          if np.isnan(df_hf.iloc[sec_line_h,5*i+1])==False:
                              b_2=b_2+1 
                          else:
                              b_2=b_2
                  f.writelines('          <Key Name="ROUGHNESS">\n\
            <Values>\n\
              <Value>%.2f</Value>\n\
            </Values>\n\
            <Unit>mm</Unit>\n\
            <DefaultUnit>m</DefaultUnit>\n\
          </Key>\n\
          <Key Name="LABEL">\n\
            <Values>\n\
              <Value>PIPE-2</Value>\n\
            </Values>\n\
            <Unit />\n\
            <DefaultUnit>NoUnit</DefaultUnit>\n\
          </Key>\n\
          <Key Name="NSEGMENT">\n\
            <Values>\n\
              <Value>%d</Value>\n\
            </Values>\n\
            <Unit />\n\
            <DefaultUnit>NoUnit</DefaultUnit>\n\
          </Key>\n\
          <Key Name="LSEGMENT">\n\
            <Values>\n'%(df.iloc[i,5],b_2))
                  for m in range(0,b_2):
                      f.writelines('              <Value>%.4f</Value>\n'%df_hf.iloc[m,5*i+1])
                  f.writelines('            </Values>\n\
            <Unit>m</Unit>\n\
            <DefaultUnit>m</DefaultUnit>\n\
          </Key>\n\
          <Key Name="LENGTH">\n\
            <Values>\n\
              <Value>%.2f</Value>\n\
            </Values>\n\
            <Unit>m</Unit>\n\
            <DefaultUnit>m</DefaultUnit>\n\
          </Key>\n\
          <Key Name="ELEVATION">\n\
            <Values>\n\
              <Value>-%.2f</Value>\n\
            </Values>\n\
            <Unit>m</Unit>\n\
            <DefaultUnit>m</DefaultUnit>\n\
          </Key>\n\
          <Key Name="DIAMETER">\n\
            <Values>\n\
              <Value>%.2f</Value>\n\
            </Values>\n\
            <Unit>mm</Unit>\n\
            <DefaultUnit>m</DefaultUnit>\n\
          </Key>\n\
        </KeyCollection>\n\
      </Keyword>\n\
      <Keyword>\n'%(df.iloc[i,7],df.iloc[i,7],df.iloc[i,11]))
                  b_3=0
                  for thir_line_h in range(0,df_hf.shape[0]):
                          if np.isnan(df_hf.iloc[thir_line_h,5*i+2])==False:
                              b_3=b_3+1 
                          else:
                              b_3=b_3
                  f.writelines('        <Tag>FLOWPATH_1.PIPE_10</Tag>\n\
        <Type>PIPE</Type>\n\
        <KeyCollection>\n\
          <Key Name="ROUGHNESS">\n\
            <Values>\n\
              <Value>%.2f</Value>\n\
            </Values>\n\
            <Unit>mm</Unit>\n\
            <DefaultUnit>m</DefaultUnit>\n\
          </Key>\n\
          <Key Name="LABEL">\n\
            <Values>\n\
              <Value>PIPE-3</Value>\n\
            </Values>\n\
            <Unit />\n\
            <DefaultUnit>NoUnit</DefaultUnit>\n\
          </Key>\n\
          <Key Name="NSEGMENT">\n\
            <Values>\n\
              <Value>%d</Value>\n\
            </Values>\n\
            <Unit />\n\
            <DefaultUnit>NoUnit</DefaultUnit>\n\
          </Key>\n\
          <Key Name="LSEGMENT">\n\
            <Values>\n'%(df.iloc[i,5],b_3))
                  for m in range(0,b_3):
                      f.writelines('              <Value>%.4f</Value>\n'%df_hf.iloc[m,5*i+2])
                  f.writelines('            </Values>\n\
            <Unit>m</Unit>\n\
            <DefaultUnit>m</DefaultUnit>\n\
          </Key>\n\
          <Key Name="LENGTH">\n\
            <Values>\n\
              <Value>%.2f</Value>\n\
            </Values>\n\
            <Unit>m</Unit>\n\
            <DefaultUnit>m</DefaultUnit>\n\
          </Key>\n\
          <Key Name="ELEVATION">\n\
            <Values>\n\
              <Value>0</Value>\n\
            </Values>\n\
            <Unit>m</Unit>\n\
            <DefaultUnit>m</DefaultUnit>\n\
          </Key>\n\
          <Key Name="DIAMETER">\n\
            <Values>\n\
              <Value>%.2f</Value>\n\
            </Values>\n\
            <Unit>mm</Unit>\n\
            <DefaultUnit>m</DefaultUnit>\n\
          </Key>\n\
        </KeyCollection>\n\
      </Keyword>\n\
      <Keyword>\n'%(df.iloc[i,8],df.iloc[i,11]))
                  b_4=0
                  for for_line_h in range(0,df_hf.shape[0]):
                          if np.isnan(df_hf.iloc[for_line_h,5*i+3])==False:
                              b_4=b_4+1 
                          else:
                              b_4=b_4
                  f.writelines('        <Tag>FLOWPATH_1.PIPE_11</Tag>\n\
        <Type>PIPE</Type>\n\
        <KeyCollection>\n\
          <Key Name="ROUGHNESS">\n\
            <Values>\n\
              <Value>%.2f</Value>\n\
            </Values>\n\
            <Unit>mm</Unit>\n\
            <DefaultUnit>m</DefaultUnit>\n\
          </Key>\n\
          <Key Name="LABEL">\n\
            <Values>\n\
              <Value>PIPE-4</Value>\n\
            </Values>\n\
            <Unit />\n\
            <DefaultUnit>NoUnit</DefaultUnit>\n\
          </Key>\n\
          <Key Name="NSEGMENT">\n\
            <Values>\n\
              <Value>%d</Value>\n\
            </Values>\n\
            <Unit />\n\
            <DefaultUnit>NoUnit</DefaultUnit>\n\
          </Key>\n\
          <Key Name="LSEGMENT">\n\
            <Values>\n'%(df.iloc[i,5],b_4))
                  for m in range(0,b_4):
                      f.writelines('              <Value>%.4f</Value>\n'%df_hf.iloc[m,5*i+3])
                  f.writelines('            </Values>\n\
            <Unit>m</Unit>\n\
            <DefaultUnit>m</DefaultUnit>\n\
          </Key>\n\
          <Key Name="LENGTH">\n\
            <Values>\n\
              <Value>%.2f</Value>\n\
            </Values>\n\
            <Unit>m</Unit>\n\
            <DefaultUnit>m</DefaultUnit>\n\
          </Key>\n\
          <Key Name="ELEVATION">\n\
            <Values>\n\
              <Value>%.2f</Value>\n\
            </Values>\n\
            <Unit>m</Unit>\n\
            <DefaultUnit>m</DefaultUnit>\n\
          </Key>\n\
          <Key Name="DIAMETER">\n\
            <Values>\n\
              <Value>%.2f</Value>\n\
            </Values>\n\
            <Unit>mm</Unit>\n\
            <DefaultUnit>m</DefaultUnit>\n\
          </Key>\n\
        </KeyCollection>\n\
      </Keyword>\n\
      <Keyword>\n'%(df.iloc[i,9],df.iloc[i,9],df.iloc[i,11]))
                  b_5=0
                  for fif_line_h in range(0,df_hf.shape[0]):
                          if np.isnan(df_hf.iloc[fif_line_h,5*i+4])==False:
                              b_5=b_5+1 
                          else:
                              b_5=b_5
                  f.writelines('        <Tag>FLOWPATH_1.PIPE_12</Tag>\n\
        <Type>PIPE</Type>\n\
        <KeyCollection>\n\
          <Key Name="ROUGHNESS">\n\
            <Values>\n\
              <Value>%.2f</Value>\n\
            </Values>\n\
            <Unit>mm</Unit>\n\
            <DefaultUnit>m</DefaultUnit>\n\
          </Key>\n\
          <Key Name="LABEL">\n\
            <Values>\n\
              <Value>PIPE-5</Value>\n\
            </Values>\n\
            <Unit />\n\
            <DefaultUnit>NoUnit</DefaultUnit>\n\
          </Key>\n\
          <Key Name="NSEGMENT">\n\
            <Values>\n\
              <Value>%d</Value>\n\
            </Values>\n\
            <Unit />\n\
            <DefaultUnit>NoUnit</DefaultUnit>\n\
          </Key>\n\
          <Key Name="LSEGMENT">\n\
            <Values>\n'%(df.iloc[i,5],b_5))
                  for m in range(0,b_5):
                      f.writelines('              <Value>%.4f</Value>\n'%df_hf.iloc[m,5*i+4])
                  f.writelines('            </Values>\n\
            <Unit>m</Unit>\n\
            <DefaultUnit>m</DefaultUnit>\n\
          </Key>\n\
          <Key Name="LENGTH">\n\
            <Values>\n\
              <Value>%.2f</Value>\n\
            </Values>\n\
            <Unit>m</Unit>\n\
            <DefaultUnit>m</DefaultUnit>\n\
          </Key>\n\
          <Key Name="ELEVATION">\n\
            <Values>\n\
              <Value>0</Value>\n\
            </Values>\n\
            <Unit>m</Unit>\n\
            <DefaultUnit>m</DefaultUnit>\n\
          </Key>\n\
          <Key Name="DIAMETER">\n\
            <Values>\n\
              <Value>%.2f</Value>\n\
            </Values>\n\
            <Unit>mm</Unit>\n\
            <DefaultUnit>m</DefaultUnit>\n\
          </Key>\n\
        </KeyCollection>\n\
      </Keyword>\n\
      <Keyword>\n'%(df.iloc[i,10],df.iloc[i,11]))
                  f.writelines('        <Tag>FLOWPATH_1.HEATTRANSFER_13</Tag>\n\
        <Type>HEATTRANSFER</Type>\n\
        <KeyCollection>\n\
          <Key Name="LABEL">\n\
            <Values>\n\
              <Value>HEATTRANS-1</Value>\n\
            </Values>\n\
            <Unit />\n\
            <DefaultUnit>NoUnit</DefaultUnit>\n\
          </Key>\n\
          <Key Name="UVALUE">\n\
            <Values>\n\
              <Value>%.2f</Value>\n\
            </Values>\n\
            <Unit>W/m2-C</Unit>\n\
            <DefaultUnit>W/m2-C</DefaultUnit>\n\
          </Key>\n\
          <Key Name="TAMBIENT">\n\
            <Values>\n\
              <Value>%.2f</Value>\n'%(df.iloc[i,12],df.iloc[i,13]))
                  f.writelines(alldata[3645:3730])
                  
             f.close
         QMessageBox.about(self, '提示', '文件已生成')            

class MyFigureCanvas(FigureCanvas):
    """
    画布
    """
    def __init__(self):
        # 画布上初始化一个图像
        self.figure = Figure()
        super().__init__(self.figure)

class child2(QMainWindow, Ui_Form3):
    def __init__(self, parent=None):
        super(child2, self).__init__(parent)
        self.setupUi(self)
        self.pushButton.clicked.connect(self.openfile_1)
        self.pushButton_3.clicked.connect(self.openfile_2)
    def open_h(self):
        self.show()
    def openfile_1(self):
        fileName=QtWidgets.QFileDialog.getExistingDirectory(None,"选取文件夹")
        self.textBrowser.insertPlainText(fileName) 
     ########   
    def openfile_2(self):
        self.textBrowser_2.clear()
        fileName,fileType=QtWidgets.QFileDialog.getOpenFileName(self,"选取文件",os.getcwd(),
                                                                "All Files(*);;Text Files(*.txt)")
    #############
    
#        self.figure = Figure()
#        super().__init__(self.figure)
    def initUI(self):
        
        self.widget = QWidget(self)
        # 设置中心窗口
        self.setCentralWidget(self.widget)
        # 创建一个网格布局管理器
        self.layout = QGridLayout()
        self.figureCanvas = MyFigureCanvas()
#        self.__draw_figure__()
        f='a.xlsx'
        dataset=pd.read_excel(f)
        
        # dataset=dataset.astype(int)
        
        # 下面的步骤和调用的方法和plot大致相同，这里我写一个简单的折线图
        self.axes = self.figureCanvas.figure.add_subplot(111)
        self.axes.set_title("line chart")
        self.axes.set_xlabel("x")
        self.axes.set_ylabel("y")
        x = dataset['A']
        y=dataset['B']
        
        # y = [15, 13, 14.5, 17, 20, 25, 26, 26, 27, 22, 18, 15]
        self.axes.plot(x, y,label='Inline label')
    
        
        
        # 工具栏 用于操作图片
        self.navigationToolbar = NavigationToolbar2QT(self.figureCanvas, self)
        self.layout.addWidget(self.navigationToolbar, 0, 0, 1, 1)
        self.layout.addWidget(self.figureCanvas, 1, 0, 1, 1)
        self.widget.setLayout(self.layout)
        self.setGeometry(300, 300, 500, 500)
        self.setWindowTitle('figure嵌入pyqt界面样例')
    def show_pic(self):
        self.show()
        
    ##########
    def generate(self):  
        import xlwt
        import os 
        import pandas as pd
        
        # base_path = r'C:\Users\zhuzhy12.GLOBAL\Desktop\olga-post-processing\文件顺序'
        def txt_xls(filename,xlsname):#####txt转excel函数
            try:
                f = open(base_path+'/'+filename)
                xls = xlwt.Workbook()
                #生成excel的方法，声明excel
                sheet = xls.add_sheet('sheet',cell_overwrite_ok=True)
                x = 0   #在excel开始写的位置（y）
         
                while True:     #循环读取文本里面的内容
                    line = f.readline()     #一行一行的读
                    if not line:    #如果没有内容，则退出循环
                        break
                    for i in range(len(line.split('\t'))):   #\t即tab健分隔
                        item = line.split('\t')[i]
                        sheet.write(x,i,item)      #x单元格经度，i单元格纬度
                    x += 1  #另起一行
                f.close()
                xls.save(base_path+'/'+xlsname)        #保存为xls文件
            except:
                raise
                      
         ####### 对tpl文件按照文件名进行排序
        time=[]
        # base_path=os.getcwd()
        base_path=self.textBrowser.toPlainText()
        files = os.listdir(base_path)
        files_txt=[]
        for path in files:
            a=os.path.splitext(path)
            if a[1]=='.tpl':
                files_txt.append(path)
                # files_txt.sort(key=lambda x: int(x.split('.')[0])) 
                files_txt.sort(key=lambda x: int(x.split('-')[-1][:-4]))        
                #####split对文件名按照'_'切片，分成test和2022.tpl两部分，取其中的数字部分进行排序
                time.append(path.split('-')[-1][:-4])
         #######   ######
        shuchu=[]
        for path in files_txt:
            a_tpl=os.path.splitext(path)#读取文件名称 
            full_path = os.path.join(base_path, path)
            # print(full_path)
            # with open(full_path) as fp:
                # data = fp.read()
                # print(data)
        ############################
            filename = '%s.txt'%a_tpl[0]
            xlsname = '%s.xlsx'%a_tpl[0]
            ######将tpl文件中相应部分写入txt文件
            f=open(base_path+'/'+'%s.tpl'%a_tpl[0],'r',encoding='utf-8')
            lines=f.readlines()
            alldata=[]
            for lines in lines:
                alldata.append(lines)
                
            
            f=open(base_path+'/'+'%s.tpl'%a_tpl[0],'r',encoding='utf-8')   
            data=[] 
            lines=f.readlines() 
            for lines in lines:
                data.append(lines)
                if 'TIME SERIES  ' in lines:
                    break   
                # while 'TIME SERIES  ' in lines:
            data_cha = [i for i in alldata if i not in data]
            # print(data.index("DATE\n"))
            ####读取标题列*****
            f=open(base_path+'/'+'%s.tpl'%a_tpl[0],'r',encoding='utf-8')   
            biaoti=[]
            lines=f.readlines()
            for lines in lines:
                if alldata.index(lines)>alldata.index('18\n'):
                    biaoti.append(lines)
            #     if alldata.index(lines)>alldata.index('18'):
                    if 'TIME SERIES' in lines:
                        break
            biaoti.pop(-1)   
            ##### ********    
              
            # data_cha=alldata-data       
              #将数据写入txt文件
            with open(base_path+'/'+'%s.txt'%a_tpl[0],'w')as f:
                for u in range(len(data_cha)):
                    f.writelines(data_cha[u])
            ######
            
            txt_xls(filename,xlsname)#将txt文件写入xls文件
            df=pd.read_excel(base_path+'/'+xlsname,header=None,names='A')
            
            df2=df['A'].str.split(" ",expand=True)###将单元格内的值分开
            # df2.astype('int64')
            df2.to_excel(base_path+'/'+'%s.xlsx'%a_tpl[0],index=False,header=None)###将分开后的数据重新写入excel 
            df2=pd.read_excel(base_path+'/'+'%s.xlsx'%a_tpl[0],header=None)
          
            biaoti=['时间']+biaoti
            for i in range(0,df2.shape[0]):
                df2.iloc[i,0]=int(df2.iloc[i,0])#将时间数据取整
            df2.columns=biaoti   
            # df2.to_excel('%s.xlsx'%a_tpl[0],index=False) #将标题写入
            ################
            for line in biaoti:
                if line=="PT 'POSITION:' 'POS-1' '(PA)' 'Pressure'\n":
                    inlet_p=biaoti.index("PT 'POSITION:' 'POS-1' '(PA)' 'Pressure'\n")
                    # print(inlet_p)
                if line=="TM 'POSITION:' 'POS-1' '(C)' 'Fluid temperature'\n":
                    inlet_t=biaoti .index("TM 'POSITION:' 'POS-1' '(C)' 'Fluid temperature'\n")
                if line=="PT 'POSITION:' 'POS-2' '(PA)' 'Pressure'\n":
                    outlet_p=biaoti .index("PT 'POSITION:' 'POS-2' '(PA)' 'Pressure'\n")
                if line=="TM 'POSITION:' 'POS-2' '(C)' 'Fluid temperature'\n":
                    outlet_t=biaoti .index("TM 'POSITION:' 'POS-2' '(C)' 'Fluid temperature'\n")
                if line=="UG 'POSITION:' 'POS-2' '(M/S)' 'Gas velocity'\n":
                    UG=biaoti .index("UG 'POSITION:' 'POS-2' '(M/S)' 'Gas velocity'\n")
                if line=="UL 'POSITION:' 'POS-2' '(M/S)' 'Average liquid film velocity'\n":
                    UL=biaoti .index("UL 'POSITION:' 'POS-2' '(M/S)' 'Average liquid film velocity'\n")
                if line=="EVR 'POSITION:' 'POS-2' '(-)' 'Erosional velocity ratio'\n":
                    EVR=biaoti .index("EVR 'POSITION:' 'POS-2' '(-)' 'Erosional velocity ratio'\n")   
                if line=="LIQC 'BRANCH:' 'FLOWPATH_1' '(M3)' 'Total liquid content in branch'\n":
                    LIQC=biaoti .index("LIQC 'BRANCH:' 'FLOWPATH_1' '(M3)' 'Total liquid content in branch'\n")      
                    # print(LIQC)
            ###############
            shuchu.append(round(((df2.iloc[-1,inlet_p]-df2.iloc[-1,outlet_p])*1.05+df2.iloc[-1,outlet_p])/1000000,2))
            shuchu.append(round(df2.iloc[-1,outlet_t],2))
            shuchu.append(round(df2.iloc[-1,UG],2))
            shuchu.append(round(df2.iloc[-1,UL],2))
            shuchu.append(round(df2.iloc[-1,EVR],2))
            shuchu.append(round(df2.iloc[-1,LIQC]*1.2,2))
            os.remove(base_path+'/'+'%s.xlsx'%a_tpl[0])
            # os.remove('%s.xls'%a_tpl[0])
            os.remove(base_path+'/'+'%s.txt'%a_tpl[0])
        df_jieguo=pd.read_excel('标准表.xlsx',header=None)    
        for i in range(0, len(shuchu),6):
            hangshu=i//6
            df_jieguo.iloc[hangshu+2,4]=shuchu[i]
        for i in range(0,len(time)):
            df_jieguo.iloc[i+2,0]=time[i]
        m=len(shuchu)//6
        for i in range(1,m+1):    
            for j in range(1+6*(i-1),6+6*(i-1)):
                hangshu=j//6
                yushu=j%6
                df_jieguo.iloc[hangshu+2,yushu+6]=shuchu[j]
                # ws.cell(row=hangshu+2,column=yushu+6).value=shuchu[j]
        # for i in range(len(time)+2,-1):
        # df2.drop(df_jieguo.index[len(time)+2:-1],inplace=False)    
        df_jieguo=df_jieguo.drop(df_jieguo.index[len(time)+2:100])
        df_jieguo.to_excel('标准表_添加结果.xlsx',index=False,header=None)
          
        QMessageBox.about(self, '提示', '文件已生成')  

  
    
if __name__ == "__main__":
    #固定的，PyQt5程序都需要QApplication对象。sys.argv是命令行参数列表，确保程序可以双击运行
    app = QApplication(sys.argv)
    main=MyMainForm()
    child=child()
    child2=child2()
    main.show()
    
        
#    main.pushButton.clicked.connect(child.open)
    main.pushButton_2.clicked.connect(main.close)
    #初始化
#    myWin = MyMainForm()
#    #将窗口控件显示在屏幕上
#    myWin.show()
    #程序运行，sys.exit方法确保程序完整退出。
    sys.exit(app.exec_())